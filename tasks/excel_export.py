"""
Excel Export Module
===================

Provides standardized Excel file formatting for Excelerant database exports.

Formatting conventions:
- Font: Times New Roman, size 9
- Header: Bold, centered, frozen row, auto-filter
- Link column: Blue hyperlinks to audio files
- Path column: Shrink-to-fit
- Gloss column: Left-aligned, italic
- All others: Centered
"""

import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# Column Ordering
# =============================================================================

# Priority columns in preferred order (before link/path columns)
PRIORITY_COLUMNS_BEFORE = [
    'index',
    'sub_index',
    'speaker',
    'token',
    'other',
    'gloss',
    'entry',
    'word',
    'P', 'R', 'C', 'M', 'V', 'F', 'T',  # Segmentation columns
    'entry_orig',
]

# For backwards compatibility
PRIORITY_COLUMNS = PRIORITY_COLUMNS_BEFORE + ['link', 'path']


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reorder DataFrame columns with priority columns first.

    Column order:
    1. Priority columns (index, sub_index, speaker, token, etc.)
    2. All link columns grouped together (link, link_*, etc.)
    3. All path columns grouped together (path, path_*, etc.)
    4. Any remaining columns

    Args:
        df: DataFrame to reorder

    Returns:
        DataFrame with reordered columns
    """
    df = df.copy()

    # Get priority columns that exist in the dataframe
    priority_cols = [col for col in PRIORITY_COLUMNS_BEFORE if col in df.columns]

    # Get all link columns (link, link_M1_1, etc.) - sorted for consistency
    link_cols = sorted([col for col in df.columns if col == 'link' or col.startswith('link_')])

    # Get all path columns (path, path_M1_1, etc.) - sorted for consistency
    path_cols = sorted([col for col in df.columns if col == 'path' or col.startswith('path_')])

    # Get remaining columns (not in any of the above)
    used_cols = set(priority_cols + link_cols + path_cols)
    remaining_cols = [col for col in df.columns if col not in used_cols]

    # Combine: priority → links → paths → remaining
    new_order = priority_cols + link_cols + path_cols + remaining_cols

    return df[new_order]


# =============================================================================
# Link Column
# =============================================================================

def add_link_column(df: pd.DataFrame, path_col: str = 'path') -> pd.DataFrame:
    """
    Add a 'link' column with 'listen' text where path exists.

    Args:
        df: DataFrame with path column
        path_col: Name of the path column

    Returns:
        DataFrame with 'link' column added
    """
    df = df.copy()

    if path_col in df.columns:
        df['link'] = df[path_col].apply(lambda x: 'listen' if pd.notna(x) else None)

    return df


# =============================================================================
# Excel Generation
# =============================================================================

def generate_excel_file(
    df: pd.DataFrame,
    path_col: str | None = 'path',
    link_col: str | None = 'link',
    gloss_col: str | None = 'gloss',
    sheet_name: str = 'Lexical Database',
    platform: str = 'windows',
) -> bytes:
    """
    Generate a formatted Excel file from DataFrame.

    Applies Excelerant standard formatting:
    - Times New Roman, size 9
    - Bold centered headers with freeze and auto-filter
    - Hyperlinked 'link' column (blue, underlined) - Windows only
    - Italic left-aligned gloss column
    - Shrink-to-fit path column
    - Centered alignment for all other columns

    Args:
        df: DataFrame to export
        path_col: Name of path column (for hyperlinks)
        link_col: Name of link column (gets hyperlinks on Windows)
        gloss_col: Name of gloss column (italic formatting)
        sheet_name: Name for the Excel sheet
        platform: 'windows' or 'mac' - controls hyperlink behavior

    Returns:
        Excel file as bytes
    """
    # Create Excel file in memory
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Load workbook for formatting
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Apply auto filter and freeze top row
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Define fonts and alignments
    font_regular = Font(name="Times New Roman", size=9)
    font_bold = Font(name="Times New Roman", size=9, bold=True)
    font_italic = Font(name="Times New Roman", size=9, italic=True)
    font_hyperlink = Font(name="Times New Roman", size=9, color="0563C1", underline="single")

    alignment_left = Alignment(horizontal='left', vertical='center')
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_shrink = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)

    # Find all link and path columns (including multi-column variants like link_M1_1)
    link_cols = [col for col in df.columns if col == 'link' or col.startswith('link_')]
    path_cols = [col for col in df.columns if col == 'path' or col.startswith('path_')]

    # Build mapping of link column -> corresponding path column
    # link_M1_1 -> path_M1_1, link -> path
    link_to_path = {}
    for link_col_name in link_cols:
        if link_col_name == 'link':
            if 'path' in path_cols:
                link_to_path[link_col_name] = 'path'
        else:
            # Extract suffix (e.g., "M1_1" from "link_M1_1")
            suffix = link_col_name[5:]  # Remove "link_"
            path_col_name = f'path_{suffix}'
            if path_col_name in path_cols:
                link_to_path[link_col_name] = path_col_name

    # Get column numbers for special formatting
    link_col_nums = set()
    path_col_nums = set()
    for col in link_cols:
        link_col_nums.add(df.columns.get_loc(col) + 1)
    for col in path_cols:
        path_col_nums.add(df.columns.get_loc(col) + 1)

    gloss_col_num = None
    if gloss_col and gloss_col in df.columns:
        gloss_col_num = df.columns.get_loc(gloss_col) + 1

    # Add hyperlinks to link columns (Windows only)
    # Mac link columns contain terminal commands, not hyperlinks
    if platform == 'windows':
        for link_col_name, path_col_name in link_to_path.items():
            link_col_idx = df.columns.get_loc(link_col_name) + 1
            path_col_idx = df.columns.get_loc(path_col_name) + 1

            for row_idx in range(2, len(df) + 2):
                path_val = ws.cell(row=row_idx, column=path_col_idx).value

                if path_val is not None:
                    link_cell = ws.cell(row=row_idx, column=link_col_idx)
                    link_cell.value = 'listen'
                    link_cell.hyperlink = str(path_val)
                    link_cell.font = font_hyperlink
                    link_cell.alignment = alignment_center

    # Apply formatting to all cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                # Skip link columns (already formatted above for Windows)
                if cell.column in link_col_nums:
                    continue
                # Format path columns with shrink to fit
                elif cell.column in path_col_nums:
                    cell.alignment = alignment_shrink
                    cell.font = font_regular
                # Format gloss column
                elif gloss_col_num and cell.column == gloss_col_num:
                    cell.alignment = alignment_left
                    cell.font = font_italic
                # Format all other columns
                else:
                    cell.alignment = alignment_center
                    cell.font = font_regular

    # Make header row bold and centered
    for cell in ws[1]:
        cell.alignment = alignment_center
        cell.font = font_bold

    # Add word column formulae if segmentation columns exist
    seg_cols = ['P', 'R', 'C', 'M', 'V', 'F', 'T']
    if 'word' in df.columns and all(col in df.columns for col in seg_cols):
        word_col_idx = df.columns.get_loc('word') + 1
        seg_letters = {
            col: get_column_letter(df.columns.get_loc(col) + 1)
            for col in seg_cols
        }
        for row_idx in range(2, len(df) + 2):
            concat = '&'.join(f'{seg_letters[col]}{row_idx}' for col in seg_cols)
            formula = f'=SUBSTITUTE({concat},"\u2205","")'
            cell = ws.cell(row=row_idx, column=word_col_idx)
            cell.value = formula
            cell.alignment = alignment_center
            cell.font = font_regular

    # Save to bytes
    formatted_output = BytesIO()
    wb.save(formatted_output)
    return formatted_output.getvalue()


def prepare_and_export(
    df: pd.DataFrame,
    path_col: str | None = 'path',
    gloss_col: str | None = 'gloss',
    add_links: bool = True,
    reorder: bool = True,
    sheet_name: str = 'Lexical Database',
    platform: str = 'windows',
) -> bytes:
    """
    Convenience function to prepare DataFrame and export to Excel.

    Combines column reordering, link column addition, and Excel generation.

    Args:
        df: DataFrame to export
        path_col: Name of path column
        gloss_col: Name of gloss column
        add_links: Whether to add link column
        reorder: Whether to reorder columns
        sheet_name: Name for the Excel sheet
        platform: 'windows' or 'mac' - controls hyperlink behavior

    Returns:
        Excel file as bytes
    """
    df = df.copy()

    # Add link column if paths exist
    if add_links and path_col and path_col in df.columns:
        df = add_link_column(df, path_col)

    # Reorder columns
    if reorder:
        df = reorder_columns(df)

    # Determine link column
    link_col = 'link' if 'link' in df.columns else None

    # Generate Excel
    return generate_excel_file(
        df=df,
        path_col=path_col,
        link_col=link_col,
        gloss_col=gloss_col,
        sheet_name=sheet_name,
        platform=platform,
    )


def export_database(df: pd.DataFrame, platform: str = 'windows') -> bytes:
    """
    Export database to formatted Excel file with automatic column detection.

    This is the primary function for exporting databases throughout Excelerant.
    It automatically detects which columns exist and applies appropriate formatting.

    All Excel exports should use this function to ensure consistent formatting.
    Update this function to change formatting across the entire application.

    Args:
        df: DataFrame to export
        platform: 'windows' or 'mac' - controls hyperlink behavior

    Returns:
        Excel file as bytes
    """
    df = df.copy()

    # Auto-detect columns
    has_path = 'path' in df.columns or any(col.startswith('path_') for col in df.columns)
    has_gloss = 'gloss' in df.columns

    # Reorder columns
    df = reorder_columns(df)

    # Determine columns for formatting
    path_col = 'path' if 'path' in df.columns else None
    gloss_col = 'gloss' if has_gloss else None
    link_col = 'link' if 'link' in df.columns else None

    # Generate Excel with standard formatting
    return generate_excel_file(
        df=df,
        path_col=path_col,
        link_col=link_col,
        gloss_col=gloss_col,
        sheet_name='Lexical Database',
        platform=platform,
    )
